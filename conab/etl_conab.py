"""
etl_conab.py
===============================================================================
Lê conab/raw/Boletim_Safra_Graos.xlsx e gera conab_atual.csv na raiz do repositório.

Detecta ano e mês automaticamente a partir do conteúdo de conab/ultimo_xlsx.txt.
Se a estrutura do XLSX mudar de forma incompatível (abas em posições erradas,
arquivo ausente, etc.), aborta com exit 1 SEM sobrescrever o CSV existente.

Uso:
    python conab/etl_conab.py    (a partir da raiz do repositório)

Dependências:
    pip install pandas openpyxl
===============================================================================
"""

import sys
import re
from pathlib import Path
import pandas as pd
import openpyxl

PASTA            = Path(__file__).resolve().parent          # conab/
PASTA_RAW        = PASTA / "raw"                            # conab/raw/
ARQUIVO_CSV      = PASTA.parent / "conab_atual.csv"         # raiz do repositório

# AJUSTE 1: Alinhando nomes dos arquivos com o script de download
ARQUIVO_XLSX     = PASTA_RAW / "Boletim_Safra_Graos.xlsx"
ARQUIVO_REGISTRO = PASTA / "ultimo_xlsx.txt"

REGIAO_MAP = {
    "AC": "Norte",        "AP": "Norte",        "AM": "Norte",   "PA": "Norte",
    "RO": "Norte",        "RR": "Norte",        "TO": "Norte",
    "AL": "Nordeste",     "BA": "Nordeste",     "CE": "Nordeste", "MA": "Nordeste",
    "PB": "Nordeste",     "PE": "Nordeste",     "PI": "Nordeste", "RN": "Nordeste",
    "SE": "Nordeste",
    "DF": "Centro-Oeste", "GO": "Centro-Oeste", "MT": "Centro-Oeste", "MS": "Centro-Oeste",
    "ES": "Sudeste",      "MG": "Sudeste",      "RJ": "Sudeste",  "SP": "Sudeste",
    "PR": "Sul",          "RS": "Sul",          "SC": "Sul",
}
UFS_VALIDAS = set(REGIAO_MAP.keys())

MESES_PT = {
    "jan": "01", "fev": "02", "mar": "03", "abr": "04",
    "mai": "05", "jun": "06", "jul": "07", "ago": "08",
    "set": "09", "out": "10", "nov": "11", "dez": "12",
}

# (indice_aba, keyword_validacao, label_csv, eh_inverno)
ABAS_CONFIG = [
    (39, "soja",   "Soja",   False),
    (38, "milho",  "Milho",  False),
    (31, "feij",   "Feijao", False),
    (44, "cevada", "Cevada", True),
    (45, "trigo",  "Trigo",  True),
]

MIN_ABAS = 46
SEP = "=" * 62


class ErroEstrutural(Exception):
    pass


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------

def parsear_versao() -> tuple[int, str]:
    """
    Lê ultimo_xlsx.txt e extrai (ano, mes_str) a partir do nome do arquivo salvo lá dentro.
    Exemplo do conteúdo: 'site_previsao_de_safra-por_produto-mai-2026.xlsx' -> (2026, '05')
    """
    if not ARQUIVO_REGISTRO.exists():
        raise ErroEstrutural(
            f"Arquivo de registro não encontrado: {ARQUIVO_REGISTRO}\n"
            "Execute baixar_conab.py primeiro."
        )
    
    # AJUSTE 2: Lendo o nome original de dentro do arquivo de texto de controle
    nome = ARQUIVO_REGISTRO.read_text(encoding="utf-8").strip()
    
    m = re.search(r'-([a-z]{3})-(\d{4})\.xlsx$', nome.lower())
    if not m:
        raise ErroEstrutural(
            f"Não foi possível extrair ano/mês do texto '{nome}' encontrado em {ARQUIVO_REGISTRO.name}.\n"
            "Padrão esperado dentro do txt: ...-MMM-AAAA.xlsx  (ex: ...-mai-2026.xlsx)"
        )
    
    mes_abrev, ano = m.group(1), int(m.group(2))
    if mes_abrev not in MESES_PT:
        raise ErroEstrutural(f"Mês desconhecido no nome do arquivo: '{mes_abrev}'")
    return ano, MESES_PT[mes_abrev]


def validar_estrutura(sheetnames: list[str]) -> None:
    """Garante que as abas esperadas estão nos índices corretos."""
    if len(sheetnames) < MIN_ABAS:
        raise ErroEstrutural(
            f"Esperadas >= {MIN_ABAS} abas, encontradas {len(sheetnames)}.\n"
            "A CONAB pode ter reorganizado o arquivo."
        )
    for idx, keyword, label, _ in ABAS_CONFIG:
        nome_norm = (
            sheetnames[idx]
            .lower()
            .replace("ã", "a")
            .replace("ê", "e")
            .replace("é", "e")
        )
        if keyword not in nome_norm:
            raise ErroEstrutural(
                f"Aba [{idx}] esperava conter '{keyword}', "
                f"encontrou '{sheetnames[idx]}'.\n"
                "Índices das abas podem ter mudado."
            )


def processar(wb, sheetnames: list[str], ano_ref: int, data_levantamento: str) -> list:
    all_records = []
    for idx, _, label, eh_inverno in ABAS_CONFIG:
        safra_inicio = ano_ref if eh_inverno else ano_ref - 1
        safra_fim    = ano_ref
        ws   = wb[sheetnames[idx]]
        rows = list(ws.iter_rows(values_only=True))
        count = 0
        for row in rows[7:]:
            uf_val = row[0] if len(row) > 0 else None
            if not isinstance(uf_val, str):
                continue
            uf = uf_val.strip().upper()
            if uf not in UFS_VALIDAS:
                continue
            area_raw   = row[2] if len(row) > 2 else None
            produt_raw = row[5] if len(row) > 5 else None
            prod_raw   = row[8] if len(row) > 8 else None
            area = round(float(area_raw   or 0) * 1000.0, 1)
            prod = round(float(prod_raw   or 0) * 1000.0, 1)
            produt = round(float(produt_raw or 0), 1)
            all_records.append({
                "uf":                  uf,
                "regiao":              REGIAO_MAP[uf],
                "cultura":             label,
                "safra_inicio":        safra_inicio,
                "safra_fim":           safra_fim,
                "ano_referencia":      ano_ref,
                "area_plantada_ha":    area,
                "producao_ton":        prod,
                "produtividade_kg_ha": produt,
                "data_levantamento":   data_levantamento,
                "fonte":               "CONAB",
            })
            count += 1
        print(f"  {label:8s}: {count:3d} linhas  <- aba [{idx}] {sheetnames[idx]!r}")
    return all_records


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------

def main() -> int:
    try:
        print(SEP)
        print("  ETL CONAB")
        print(SEP)

        ano_ref, mes_str = parsear_versao()
        data_levantamento = f"{ano_ref}-{mes_str}"
        print(f"\n  Boletim : {mes_str}/{ano_ref}  (data_levantamento={data_levantamento})")
        print(f"  Safra verão  : {ano_ref - 1}/{ano_ref}  (Soja, Milho, Feijão)")
        print(f"  Safra inverno: {ano_ref}/{ano_ref}  (Cevada, Trigo)")

        if not ARQUIVO_XLSX.exists():
            raise ErroEstrutural(f"XLSX não encontrado: {ARQUIVO_XLSX}")

        wb = openpyxl.load_workbook(ARQUIVO_XLSX, read_only=True, data_only=True)
        sheetnames = wb.sheetnames
        print(f"\n  Total de abas no arquivo: {len(sheetnames)}")

        validar_estrutura(sheetnames)
        print("  Estrutura validada OK\n")

        all_records = processar(wb, sheetnames, ano_ref, data_levantamento)
        wb.close()

        if not all_records:
            raise ErroEstrutural("Nenhum registro extraído — verifique as linhas de dados.")

        df = pd.DataFrame(all_records, columns=[
            "uf", "regiao", "cultura", "safra_inicio", "safra_fim",
            "ano_referencia", "area_plantada_ha", "producao_ton",
            "produtividade_kg_ha", "data_levantamento", "fonte",
        ])
        df["safra_inicio"]        = df["safra_inicio"].astype(int)
        df["safra_fim"]           = df["safra_fim"].astype(int)
        df["ano_referencia"]      = df["ano_referencia"].astype(int)
        df["area_plantada_ha"]    = df["area_plantada_ha"].astype(float)
        df["producao_ton"]        = df["producao_ton"].astype(float)
        df["produtividade_kg_ha"] = df["produtividade_kg_ha"].astype(float)
        df["cultura"] = df["cultura"].str.replace("Feijao", "Feijão", regex=False)

        df.to_csv(ARQUIVO_CSV, index=False, encoding="utf-8-sig", float_format="%.1f")

        soja_total = df[df["cultura"] == "Soja"]["producao_ton"].sum()
        print(f"\n  -> CSV salvo: {ARQUIVO_CSV}  ({len(df)} lines)")
        print(f"  -> Produção Soja total: {soja_total / 1e6:.2f} M t")
        print(f"\n[OK] CSV gerado com sucesso.")
        return 0

    except ErroEstrutural as e:
        print(f"\n[ERRO ESTRUTURAL]\n{e}")
        print("\nCSV atual mantido sem alterações.")
        return 1
    except Exception as e:
        print(f"\n[ERRO INESPERADO] {type(e).__name__}: {e}")
        print("\nCSV atual mantido sem alterações.")
        return 1


if __name__ == "__main__":
    sys.exit(main())